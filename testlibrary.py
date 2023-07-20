##coder-Nguyễn Đức Nguyên
#ver 1.0.1
import customtkinter as ctk
import tkinter,webbrowser
from tkinter import *
from tkinter import messagebox
from datetime import date
from PIL import Image,ImageTk
import os,pathlib,openpyxl,xlrd
from openpyxl import Workbook
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")
back='#06283D'
framebg='#EDEDED'
framefg='#06283DA'
colortext='white'
root=ctk.CTk()
root.iconbitmap('D:\quản lý thư viện\.ico\logo.ico')
root.title('Quản Lý Người Mượn Sách')
root.configure(bg=back)
root.geometry('1190x610+210+100')
root.resizable(False,False)
file=pathlib.Path('DANH_SACH_MUON.xlsx')
ds_admin=['Nguyễn Đức Nguyên','Nguyễn Cảnh Huy','Tô Thị Linh','ADMIN']
def seach_data(event):
    webbrowser.open("https://facebook.com/nguyen30112007")
def re_no():
    file=openpyxl.load_workbook('DANH_SACH_MUON.xlsx')
    sheet=file.active
    row=sheet.max_row
    max_row_value=sheet.cell(row=row,column=1).value
    try:
        Res.set(max_row_value+1)
    except:
        Res.set('1')
def clear():
    Name.set('')
    Lop.set('')
    sex.set('')
    place.set('')
    contact.set('')
    ts.set('')
    han.set('')
    date_muon.set('')
    re_no()
def save():
    R1=Name.get()
    N1=Lop.get()
    C1=sex.get()
    A1=place.get()
    A2=contact.get()
    A3=ts.get()
    A4=han.get()
    A5=date_muon.get()
    selected_value = b1.get()
    if R1=='':
        messagebox.showerror('Error','Tên Không Thể Để Trống')
    elif N1=='':
        messagebox.showerror('Error','Lớp Không Thể Để Trống')
    elif C1=='':
        messagebox.showerror('Error','Giới Tính Không Thể Để Trống')
    elif A1=='':
        messagebox.showerror('Error','Địa Chỉ Không Thể Để Trống')
    elif A2=='':
        messagebox.showerror('Error','Thông Tin Liên Hệ Không Thể Để Trống')
    elif A3=='':
        messagebox.showerror('Error','Tên Sách Không Thể Để Trống')
    elif A4=='':
        messagebox.showerror('Error','Ngày Mượn Không Thể Để Trống')
    elif A5=='':
        messagebox.showerror('Error','Hạn Không Thể Để Trống')
    elif selected_value=='':
        messagebox.showerror('Error','Người Duyệt Không Thể Để Trống')
    

if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='Đăng ký số:'
    sheet['B1']='Tên'
    sheet['C1']='Lớp'
    sheet['D1']='Giới Tính'
    sheet['E1']='Ngày Đăng Ký'
    sheet['F1']='Địa chỉ'
    sheet['G1']='Thông Tin Liên Lạc'
    sheet['H1']='Tên Sách Mươn'
    sheet['I1']='Mượn Ngày:'
    sheet['J1']='Ngày Trả'
    sheet['K1']='Trách Nhiệm'
    sheet['L1']='Người Xét Duyệt'
    file.save('DANH_SACH_MUON.xlsx')
link_label=ctk.CTkLabel(root,text='thông tin liên lạc:ostmay37@gmail.com',width=10,height=3,anchor='e', cursor="hand2")
link_label.pack_configure(side='top',fill='x')
link_label.bind("<Button-1>", seach_data)
qlms=ctk.CTkLabel(root,text='Quản Lý Mượn Sách',width=10,height=2,cursor="hand2")
qlms.pack_configure(side='top',fill='x')
qlms.bind("<Button-2>")
search=ctk.StringVar()
#timg kiếm
a=ctk.CTkEntry(master=root,placeholder_text="Tìm Kiếm",textvariable=search)
a.place(x=720,y=50)
se_bt=ctk.CTkButton(master=root,text='Tìm kiếm',compound=LEFT,width=123,command=lambda:messagebox.showinfo('Thông Tin','Dữ Liệu Đang Được Cập Nhật'))
se_bt.place(x=870,y=50)
udt=ctk.CTkButton(master=root,text='Cập Nhật Thông Tin',command=lambda:messagebox.showinfo('Thông Tin','Dữ Liệu Đang Được Cập Nhật'))
udt.place(x=1000, y = 50 )
ctk.CTkLabel(root,text='Danh sách mượn số:').place(x=20,y=100/2)
Res=ctk.StringVar()
Date=ctk.StringVar()
reg=ctk.CTkEntry(root,placeholder_text='số',textvariable=Res)
reg.place(x=160,y=100/2)
today=date.today()
d1=today.strftime("%d/%m/%y")
date_et=ctk.CTkEntry(root,textvariable=Date)
date_et.place(x=450,y=100/2)
Date.set(d1)
#1
lb_frame=ctk.CTkFrame(root,width=850,height=400)
lb_frame.place(x=20,y=180)
ten=ctk.CTkLabel(lb_frame,text='Tên',text_color=colortext)
ten.place(x=30,y=10)
lop=ctk.CTkLabel(lb_frame,text='Lớp',text_color=colortext)
lop.place(x=30,y=70)
gioitinh=ctk.CTkLabel(lb_frame,text='Giới Tính',text_color=colortext)
gioitinh.place(x=30,y=130)
place_lb=ctk.CTkLabel(lb_frame,text='Địa chỉ',text_color=colortext)
place_lb.place(x=30,y=190)
cto=ctk.CTkLabel(lb_frame,text='Thông Tin Liên Lạc',text_color=colortext)
cto.place(x=30,y=250)
#2
book_name=ctk.CTkLabel(lb_frame,text='Tên Sách',text_color=colortext)
book_name.place(x=450,y=10)
ngay_muon=ctk.CTkLabel(lb_frame,text='Ngày Mượn',text_color=colortext)
ngay_muon.place(x=450,y=70)
han1=ctk.CTkLabel(lb_frame,text='Hạn',text_color=colortext)
han1.place(x=450,y=130)
luuy=ctk.CTkLabel(lb_frame,text='Trách Nhiệm Nếu Hư Sách',text_color=colortext)
luuy.place(x=450,y=190)
admin_duyet=ctk.CTkLabel(lb_frame,text='Người Duyệt',text_color=colortext)
admin_duyet.place(x=450,y=250)
#entry()
Name=ctk.StringVar()
Lop=ctk.StringVar()
sex=ctk.StringVar()
place=ctk.StringVar()
contact=ctk.StringVar()
ctk.CTkEntry(lb_frame,placeholder_text='Tên',textvariable=Name).place(x=180,y=10)
qw=ctk.CTkOptionMenu(lb_frame,values=['Nam','Nữ'])
qw.place(x=180,y=70)
qw.set('Chọn')
ctk.CTkEntry(lb_frame,placeholder_text='Giới Tính',textvariable=sex).place(x=180,y=130)
ctk.CTkEntry(lb_frame,placeholder_text='Địa Chỉ',textvariable=place).place(x=180,y=190)
ctk.CTkEntry(lb_frame,placeholder_text='CONTACT',textvariable=contact).place(x=180,y=250)
#entry2
ts=ctk.StringVar()
date_muon=ctk.StringVar()
han=ctk.StringVar()
c6=ctk.CTkEntry(lb_frame,placeholder_text='Tên Sách',textvariable=ts).place(x=640,y=10)
c7=ctk.CTkEntry(lb_frame,placeholder_text='Ngày Mượn',textvariable=date_muon).place(x=640,y=70)
c8=ctk.CTkEntry(lb_frame,placeholder_text='Hạn',textvariable=han).place(x=640,y=130)
c9=ctk.CTkSegmentedButton(lb_frame,values=['Chấp Nhận','Bỏ qua']).place(x=640,y=190)
b1=ctk.CTkOptionMenu(lb_frame,values=ds_admin)
b1.place(x=640,y=250)
b1.set('Chọn Người Xét')
#Nút
nguyen=None
def admin():
    global nguyen
    if nguyen is None or not nguyen.winfo_exists() or nguyen.winfo_toplevel() != nguyen:
        nguyen=Toplevel(master=root)
        nguyen.title('Admin')
        nguyen.geometry('400x500')
        nguyen.transient(master=root)

        nguyen.grab_set()
        nguyen.mainloop()
    else:
        messagebox.showinfo('Thông Tin','Cửa sổ Admin đang được hiển thị')
image = Image.open("D:\quản lý thư viện\pngfile\\ad.png").resize((150,150),Image.LANCZOS)
photo = ImageTk.PhotoImage(image)
canvas=Button(root,image=photo,command=admin)
canvas.place(x=1000,y=160)
ctk.CTkButton(root,text='Lưu Người Mượn',command=save).place(x=1000,y=340)
ctk.CTkButton(root,text='Tải Lên',command=lambda:messagebox.showinfo('Thông Tin','Dữ Liệu Đang Được Cập Nhật')).place(x=1000,y=390)
ctk.CTkButton(root,text='Xoá Hạn Người Mượn',command=lambda:messagebox.showinfo('Thông Tin','Dữ Liệu Đang Được Cập Nhật')).place(x=1000,y=440)
ctk.CTkButton(root,text='Làm mới Bảng',command=lambda:messagebox.showinfo('Thông tin','Dữ Liệu Đang Được Cập Nhật')).place(x=1000,y=490)
ctk.CTkRadioButton(root,text='Sáng').place(x=1000,y=540)
ctk.CTkRadioButton(root,text='Tối').place(x=1000,y=540)
def check_color():
    a=radio_var.get()
    global colortext
    if a==2:
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")
        colortext='white'
        ten.configure(text_color=colortext)
        lop.configure(text_color=colortext)
        gioitinh.configure(text_color=colortext)
        place_lb.configure(text_color=colortext)
        cto.configure(text_color=colortext)
        han1.configure(text_color=colortext)
        admin_duyet.configure(text_color=colortext)
        luuy.configure(text_color=colortext)
        book_name.configure(text_color=colortext)
        ngay_muon.configure(text_color=colortext)
        cto.configure(text_color=colortext)
    elif a==1:
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        colortext='black'
        ten.configure(text_color=colortext)
        lop.configure(text_color=colortext)
        gioitinh.configure(text_color=colortext)
        place_lb.configure(text_color=colortext)
        cto.configure(text_color=colortext)
        han1.configure(text_color=colortext)
        admin_duyet.configure(text_color=colortext)
        luuy.configure(text_color=colortext)
        book_name.configure(text_color=colortext)
        ngay_muon.configure(text_color=colortext)
        cto.configure(text_color=colortext)
radio_var= tkinter.IntVar(value=0)
radiobutton_1 = ctk.CTkRadioButton(root, text='Sáng',
                                             command=check_color, variable= radio_var, value=1).place(x=1000,y=540)
radiobutton_2 = ctk.CTkRadioButton(root, text="Tối",
                                             command=check_color, variable= radio_var, value=2).place(x=1100,y=540)
root.mainloop()


